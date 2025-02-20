"""
Module for handling report export functionality in different formats.
"""
import csv
import io
import json
from typing import Any, Dict
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

logger = logging.getLogger(__name__)

def generate_excel_report(report_data: Dict[str, Any]) -> io.BytesIO:
    """Generate an Excel report from the billing data"""
    try:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Report"

        # Add headers
        headers = ["Order ID", "Service Name", "Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")
            cell.alignment = Alignment(horizontal="center")

        # Add data
        row = 2
        for order in report_data.get('orders', []):
            for service in order.get('services', []):
                ws.cell(row=row, column=1, value=order['order_id'])
                ws.cell(row=row, column=2, value=service['service_name'])
                ws.cell(row=row, column=3, value=float(service['amount']))
                row += 1

        # Add totals
        row += 1
        ws.cell(row=row, column=1, value="Total")
        ws.cell(row=row, column=3, value=float(report_data.get('total_amount', 0)))
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3).font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise

def generate_pdf_report(report_data: Dict[str, Any]) -> io.BytesIO:
    """Generate a PDF report from the billing data"""
    try:
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Add title
        elements.append(Paragraph("Billing Report", styles['Title']))
        elements.append(Paragraph(f"Customer: {report_data.get('customer_name', 'N/A')}", styles['Normal']))
        elements.append(Paragraph(f"Period: {report_data.get('start_date', '')} to {report_data.get('end_date', '')}", styles['Normal']))

        # Prepare table data
        table_data = [["Order ID", "Service Name", "Amount"]]
        for order in report_data.get('orders', []):
            for service in order.get('services', []):
                table_data.append([
                    str(order['order_id']),
                    service['service_name'],
                    f"${float(service['amount']):.2f}"
                ])

        # Add total row
        table_data.append(["Total", "", f"${float(report_data.get('total_amount', 0)):.2f}"])

        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating PDF report: {str(e)}")
        raise

def generate_csv_report(report_data: Dict[str, Any]) -> io.StringIO:
    """Generate a CSV report from the billing data"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)

        # Write headers
        writer.writerow(["Order ID", "Service Name", "Amount"])

        # Write data
        for order in report_data.get('orders', []):
            for service in order.get('services', []):
                writer.writerow([
                    order['order_id'],
                    service['service_name'],
                    f"{float(service['amount']):.2f}"
                ])

        # Write total
        writer.writerow(["Total", "", f"{float(report_data.get('total_amount', 0)):.2f}"])

        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating CSV report: {str(e)}")
        raise 