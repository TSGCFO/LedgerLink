import unittest
from unittest.mock import patch, MagicMock, call
import io
import csv
import json
from datetime import datetime
from decimal import Decimal
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate

from billing.exporters import (
    generate_excel_report,
    generate_pdf_report,
    generate_csv_report
)

class ExporterTests(unittest.TestCase):
    """Tests for the report export functionality"""
    
    def setUp(self):
        # Sample report data for testing
        self.report_data = {
            'customer_name': 'Test Customer',
            'start_date': '2023-01-01',
            'end_date': '2023-01-31',
            'orders': [
                {
                    'order_id': 'ORD-001',
                    'services': [
                        {
                            'service_id': 1,
                            'service_name': 'Shipping',
                            'amount': '10.50'
                        },
                        {
                            'service_id': 2,
                            'service_name': 'Handling',
                            'amount': '5.25'
                        }
                    ],
                    'total_amount': '15.75'
                },
                {
                    'order_id': 'ORD-002',
                    'services': [
                        {
                            'service_id': 1,
                            'service_name': 'Shipping',
                            'amount': '8.30'
                        }
                    ],
                    'total_amount': '8.30'
                }
            ],
            'total_amount': '24.05'
        }
    
    @patch('billing.exporters.Workbook')
    def test_generate_excel_report(self, mock_workbook_class):
        """Test generating an Excel report"""
        # Setup mocks
        mock_workbook = MagicMock()
        mock_worksheet = MagicMock()
        mock_cell = MagicMock()
        mock_column = MagicMock()
        
        mock_workbook_class.return_value = mock_workbook
        mock_workbook.active = mock_worksheet
        mock_worksheet.cell.return_value = mock_cell
        mock_worksheet.columns = [[mock_column]]
        mock_column.column_letter = 'A'
        
        # Expected calls for writing data
        expected_cell_calls = [
            # Headers
            call(row=1, column=1, value="Order ID"),
            call(row=1, column=2, value="Service Name"),
            call(row=1, column=3, value="Amount"),
            
            # Order 1, Service 1
            call(row=2, column=1, value="ORD-001"),
            call(row=2, column=2, value="Shipping"),
            call(row=2, column=3, value=10.5),
            
            # Order 1, Service 2
            call(row=3, column=1, value="ORD-001"),
            call(row=3, column=2, value="Handling"),
            call(row=3, column=3, value=5.25),
            
            # Order 2, Service 1
            call(row=4, column=1, value="ORD-002"),
            call(row=4, column=2, value="Shipping"),
            call(row=4, column=3, value=8.3),
            
            # Total row
            call(row=6, column=1, value="Total"),
            call(row=6, column=3, value=24.05),
        ]
        
        # Call the function
        result = generate_excel_report(self.report_data)
        
        # Verify workbook was created
        mock_workbook_class.assert_called_once()
        
        # Verify worksheet was set up
        self.assertEqual(mock_worksheet.title, "Billing Report")
        
        # Verify data was written (spot check some key calls)
        mock_worksheet.cell.assert_has_calls(expected_cell_calls, any_order=True)
        
        # Verify workbook was saved
        mock_workbook.save.assert_called_once()
        
        # Verify result is a BytesIO object
        self.assertIsInstance(result, io.BytesIO)
    
    @patch('billing.exporters.SimpleDocTemplate')
    def test_generate_pdf_report(self, mock_simpledoctemplate):
        """Test generating a PDF report"""
        # Setup mocks
        mock_doc = MagicMock()
        mock_simpledoctemplate.return_value = mock_doc
        
        # Call the function
        result = generate_pdf_report(self.report_data)
        
        # Verify SimpleDocTemplate was created with the right parameters
        mock_simpledoctemplate.assert_called_once()
        args, kwargs = mock_simpledoctemplate.call_args
        self.assertIsInstance(args[0], io.BytesIO)
        self.assertEqual(kwargs['pagesize'], letter)
        
        # Verify document was built
        mock_doc.build.assert_called_once()
        
        # Verify result is a BytesIO object
        self.assertIsInstance(result, io.BytesIO)
    
    def test_generate_csv_report(self):
        """Test generating a CSV report"""
        # Call the function
        result = generate_csv_report(self.report_data)
        
        # Reset the position to the beginning of the file
        result.seek(0)
        
        # Read the CSV content
        reader = csv.reader(result)
        rows = list(reader)
        
        # Verify headers
        self.assertEqual(rows[0], ["Order ID", "Service Name", "Amount"])
        
        # Verify data rows
        self.assertEqual(rows[1], ["ORD-001", "Shipping", "10.50"])
        self.assertEqual(rows[2], ["ORD-001", "Handling", "5.25"])
        self.assertEqual(rows[3], ["ORD-002", "Shipping", "8.30"])
        
        # Verify total row
        self.assertEqual(rows[4], ["Total", "", "24.05"])
        
        # Verify result is a StringIO object
        self.assertIsInstance(result, io.StringIO)
    
    def test_excel_report_integration(self):
        """Test generating an Excel report (integration test)"""
        result = generate_excel_report(self.report_data)
        
        # Verify the result is a BytesIO object
        self.assertIsInstance(result, io.BytesIO)
        
        try:
            # Try to load the Excel file
            wb = openpyxl.load_workbook(result)
            ws = wb.active
            
            # Verify the worksheet title
            self.assertEqual(ws.title, "Billing Report")
            
            # Verify headers
            self.assertEqual(ws.cell(row=1, column=1).value, "Order ID")
            self.assertEqual(ws.cell(row=1, column=2).value, "Service Name")
            self.assertEqual(ws.cell(row=1, column=3).value, "Amount")
            
            # Verify some data rows
            self.assertEqual(ws.cell(row=2, column=1).value, "ORD-001")
            self.assertEqual(ws.cell(row=2, column=2).value, "Shipping")
            self.assertAlmostEqual(ws.cell(row=2, column=3).value, 10.5)
            
        except Exception as e:
            self.fail(f"Failed to load Excel file: {str(e)}")
    
    def test_csv_report_integration(self):
        """Test generating a CSV report (integration test)"""
        result = generate_csv_report(self.report_data)
        
        # Verify the result is a StringIO object
        self.assertIsInstance(result, io.StringIO)
        
        # Reset the position to the beginning of the file
        result.seek(0)
        
        # Read the CSV content
        content = result.read()
        
        # Verify the CSV contains expected data
        self.assertIn("Order ID,Service Name,Amount", content)
        self.assertIn("ORD-001,Shipping,10.50", content)
        self.assertIn("ORD-001,Handling,5.25", content)
        self.assertIn("ORD-002,Shipping,8.30", content)
        self.assertIn("Total,,24.05", content)


if __name__ == '__main__':
    unittest.main()