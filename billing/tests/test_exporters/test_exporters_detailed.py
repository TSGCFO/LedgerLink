import pytest
import io
from decimal import Decimal
import json
from datetime import datetime, timedelta
import csv

from django.test import TestCase
from openpyxl import load_workbook

from billing.exporters import generate_excel_report, generate_csv_report, generate_pdf_report


class ExporterDetailedTests(TestCase):
    """Detailed tests for the billing exporters."""
    
    def setUp(self):
        """Set up test data for exporters."""
        self.sample_report_data = {
            "customer_name": "Test Company",
            "start_date": "2025-01-01",
            "end_date": "2025-01-31",
            "orders": [
                {
                    "order_id": "ORD-001",
                    "services": [
                        {
                            "service_id": 1,
                            "service_name": "Standard Shipping",
                            "amount": "25.50"
                        },
                        {
                            "service_id": 2,
                            "service_name": "Handling Fee",
                            "amount": "10.00"
                        }
                    ],
                    "total_amount": "35.50"
                },
                {
                    "order_id": "ORD-002",
                    "services": [
                        {
                            "service_id": 1,
                            "service_name": "Standard Shipping",
                            "amount": "15.75"
                        }
                    ],
                    "total_amount": "15.75"
                }
            ],
            "service_totals": {
                "1": {
                    "name": "Standard Shipping",
                    "amount": "41.25"
                },
                "2": {
                    "name": "Handling Fee",
                    "amount": "10.00"
                }
            },
            "total_amount": "51.25"
        }
    
    def test_excel_export_content(self):
        """Test that the Excel export contains the expected content."""
        output = generate_excel_report(self.sample_report_data)
        
        # Load the Excel workbook
        wb = load_workbook(output)
        ws = wb.active
        
        # Check title and headers
        self.assertEqual(ws.title, "Billing Report")
        self.assertEqual(ws.cell(row=1, column=1).value, "Order ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Service Name")
        self.assertEqual(ws.cell(row=1, column=3).value, "Amount")
        
        # Check data
        self.assertEqual(ws.cell(row=2, column=1).value, "ORD-001")
        self.assertEqual(ws.cell(row=2, column=2).value, "Standard Shipping")
        self.assertEqual(ws.cell(row=2, column=3).value, 25.50)
        
        self.assertEqual(ws.cell(row=3, column=1).value, "ORD-001")
        self.assertEqual(ws.cell(row=3, column=2).value, "Handling Fee")
        self.assertEqual(ws.cell(row=3, column=3).value, 10.00)
        
        self.assertEqual(ws.cell(row=4, column=1).value, "ORD-002")
        self.assertEqual(ws.cell(row=4, column=2).value, "Standard Shipping")
        self.assertEqual(ws.cell(row=4, column=3).value, 15.75)
        
        # Check total
        total_row = 5  # Assuming total row is after all data rows
        self.assertEqual(ws.cell(row=total_row, column=1).value, "Total")
        self.assertEqual(ws.cell(row=total_row, column=3).value, 51.25)
    
    def test_excel_export_empty_data(self):
        """Test Excel export with empty data."""
        empty_data = {
            "orders": [],
            "total_amount": "0.00"
        }
        
        output = generate_excel_report(empty_data)
        
        # Load the Excel workbook
        wb = load_workbook(output)
        ws = wb.active
        
        # Check headers exist but no data rows
        self.assertEqual(ws.cell(row=1, column=1).value, "Order ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Service Name")
        self.assertEqual(ws.cell(row=1, column=3).value, "Amount")
        
        # Check total with zero amount
        self.assertEqual(ws.cell(row=2, column=1).value, "Total")
        self.assertEqual(ws.cell(row=2, column=3).value, 0.0)
    
    def test_csv_export_content(self):
        """Test that the CSV export contains the expected content."""
        output = generate_csv_report(self.sample_report_data)
        
        # Read the CSV content
        output.seek(0)
        reader = csv.reader(output)
        rows = list(reader)
        
        # Check headers
        self.assertEqual(rows[0], ["Order ID", "Service Name", "Amount"])
        
        # Check data
        self.assertEqual(rows[1], ["ORD-001", "Standard Shipping", "25.50"])
        self.assertEqual(rows[2], ["ORD-001", "Handling Fee", "10.00"])
        self.assertEqual(rows[3], ["ORD-002", "Standard Shipping", "15.75"])
        
        # Check total
        self.assertEqual(rows[4], ["Total", "", "51.25"])
    
    def test_csv_export_empty_data(self):
        """Test CSV export with empty data."""
        empty_data = {
            "orders": [],
            "total_amount": "0.00"
        }
        
        output = generate_csv_report(empty_data)
        
        # Read the CSV content
        output.seek(0)
        reader = csv.reader(output)
        rows = list(reader)
        
        # Check headers exist but only total row
        self.assertEqual(rows[0], ["Order ID", "Service Name", "Amount"])
        self.assertEqual(rows[1], ["Total", "", "0.00"])
    
    def test_pdf_export_generation(self):
        """Test that PDF export generates a valid PDF."""
        output = generate_pdf_report(self.sample_report_data)
        
        # Check if output is a BytesIO object
        self.assertIsInstance(output, io.BytesIO)
        
        # Check if the output contains PDF signature
        output.seek(0)
        content = output.read()
        self.assertTrue(content.startswith(b'%PDF'))
    
    def test_excel_export_with_many_orders(self):
        """Test Excel export with many orders."""
        # Create a report with many orders
        large_report = {
            "orders": [],
            "total_amount": "0.00"
        }
        
        # Add 50 orders
        total_amount = Decimal("0.00")
        for i in range(1, 51):
            order_amount = Decimal(str(i * 10))
            large_report["orders"].append({
                "order_id": f"ORD-{i:03d}",
                "services": [
                    {
                        "service_id": 1,
                        "service_name": "Standard Shipping",
                        "amount": str(order_amount)
                    }
                ],
                "total_amount": str(order_amount)
            })
            total_amount += order_amount
        
        large_report["total_amount"] = str(total_amount)
        
        output = generate_excel_report(large_report)
        
        # Load the Excel workbook
        wb = load_workbook(output)
        ws = wb.active
        
        # Check the total number of rows (50 orders + header + total)
        self.assertEqual(ws.max_row, 52)
        
        # Check the total amount (sum of 10 * i for i in range(1, 51))
        self.assertEqual(ws.cell(row=52, column=3).value, float(total_amount))
    
    def test_pdf_export_with_special_characters(self):
        """Test PDF export with special characters."""
        # Create a report with special characters
        report_with_special_chars = {
            "customer_name": "Testing & Co. (Pty) Ltd.",
            "start_date": "2025-01-01",
            "end_date": "2025-01-31",
            "orders": [
                {
                    "order_id": "ORD-SP#1",
                    "services": [
                        {
                            "service_id": 1,
                            "service_name": "Premium Service™",
                            "amount": "99.99"
                        }
                    ],
                    "total_amount": "99.99"
                }
            ],
            "total_amount": "99.99"
        }
        
        output = generate_pdf_report(report_with_special_chars)
        
        # Check if output is a BytesIO object
        self.assertIsInstance(output, io.BytesIO)
        
        # Check if the output contains PDF signature
        output.seek(0)
        content = output.read()
        self.assertTrue(content.startswith(b'%PDF'))